from django.db.models.deletion import ProtectedError
from ocds_master_tables.models import (
    Address,
    Budget,
    Classification,
    ContactPoint,
    Entity,
    Identifier,
    Milestone,
    Period,
    Projet,
    Unit,
    Value,
)
from ocds_planning.models import Planning
from ocds_release.models import Record, Release, Target
from ocds_tender.models import Tender
from ocds_awards.models import Award
from ocds_contracts.models import Contract
from ocds_implementation.models import Implementation, Transaction

from openpyxl.worksheet import worksheet

def match_values(db_instance, local_instance):
    if not type(db_instance) == type(local_instance):
        raise Exception('Data instances must be instance of the same model')
    for (field_name, value) in local_instance.__dict__.items():
        if not (field_name == '_state' or field_name == 'id'):
            old_value = getattr(db_instance, field_name)
            if str(old_value) != str(value):
                setattr(db_instance, field_name, value)
                db_instance.save()

def set_related_fields(db_instance, related_fields: list):
    """
    fields : List of tuple (field_name, incoming_object)
    """
    for field in related_fields:
        if not getattr(db_instance, field[0]):
            field[1].save()
            setattr(db_instance, field[0], field[1])
        else:
            match_values(getattr(db_instance, field[0]), field[1])
    db_instance.save()

def get_engaged_objects(model, ws, col_index, lookup_field, select_related_fields=(), prefetch_related_fields=()):
    engaged_objects = {}
    for col in ws.iter_cols(min_row=4, min_col=col_index, max_col=col_index, values_only=True):
        for identifier in col:
            if not identifier:
                continue
            try:
                engaged_objects[identifier]
                continue
            except:
                try:
                    objs = model.objects
                    if select_related_fields:
                        objs = objs.select_related(select_related_fields)
                    if prefetch_related_fields:
                        objs = objs.prefetch_related(prefetch_related_fields)
                    engaged_objects[identifier] = objs.get(**{lookup_field:identifier})
                except model.DoesNotExist:
                    print("%s object with %s : %s does not exist" % (str(model), lookup_field, identifier))
                    continue
    return engaged_objects

def get_multiple_engaged_objects(ws, col_index, lookup_field, modifiers, modifier_col_index, select_related_fields=(), prefetch_related_fields=()):
    """
    modifiers : {modifier_name: modifier_class}
    """
    engaged_objects = {}
    for key, value in modifiers.items():
        engaged_objects[key] = {}

    for col in ws.iter_cols(min_row=4, min_col=col_index, max_col=col_index, values_only=True):
        for i in range(len(col)):
            identifier = col[i]
            if not identifier:
                continue
            engaged_objects_modifier = ws.cell(row=i+4, column=modifier_col_index).value
            model = modifiers[engaged_objects_modifier]
            try:
                engaged_objects[engaged_objects_modifier][identifier]
                continue
            except:
                try:
                    objs = model.objects
                    if select_related_fields:
                        objs = objs.select_related(select_related_fields)
                    if prefetch_related_fields:
                        objs = objs.prefetch_related(prefetch_related_fields)
                    engaged_objects[engaged_objects_modifier][identifier] = objs.get(**{lookup_field:identifier})
                except model.DoesNotExist:
                    print("%s object with %s : %s does not exist" % (str(model), lookup_field, identifier))
                    continue
    return engaged_objects

def create_records(ws: worksheet):
    key_map = {
        "ocid": 0,
        "target": 1,
        "address_country": 2,
        "address_region": 3,
        "address_locality": 4,
        "address_postalcode": 5,
        "address_longitude": 6,
        "address_latitude": 7
    }
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        incoming_address = Address(
            country_name=flat_object[key_map["address_country"]],
            region=flat_object[key_map["address_region"]],
            locality=flat_object[key_map["address_locality"]],
            postal_code=flat_object[key_map["address_postalcode"]],
            locality_longitude=flat_object[key_map["address_longitude"]],
            locality_latitude=flat_object[key_map["address_latitude"]]
        )
        incoming_target = Target(name=flat_object[key_map["target"]])
        try:
            record = Record.objects.select_related('implementation_address', 'target', 'compiled_release').get(ocid=flat_object[key_map["ocid"]])
            related_fields = [
                ('implementation_address', incoming_address),
                ('target', incoming_target),
            ]
            set_related_fields(record, related_fields)
        except Record.DoesNotExist:
            incoming_address.save()
            incoming_target.save()
            record = Record.objects.create(
                ocid=flat_object[key_map["ocid"]],
                implementation_address=incoming_address,
                target=incoming_target
            )
        try:
            record.compiled_release
        except Release.DoesNotExist:
            Release.objects.create(
                ref_record=record,
                tag=["planning"]
            )

def create_parties(ws: worksheet):
    key_map = {
        "party_id": 0,
        "record_ocid": 1,
        "name": 2,
        "id_schema": 3,
        "id_legal_name": 4,
        "id_uri": 5,
        "address_country": 6,
        "address_region": 7,
        "address_locality": 8,
        "address_postalcode": 9,
        "address_longitude": 10,
        "address_latitude": 11,
        "contact_name": 12,
        "contact_mail": 13,
        "contact_phone": 14,
        "contact_fax": 15,
        "contact_url": 16
    }
    releases = get_engaged_objects(Release, ws, 2, 'ref_record__ocid')
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            release = releases[flat_object[key_map["record_ocid"]]]
        except:
            print("Release reference not found")
            continue
        incoming_address = Address(
            country_name=flat_object[key_map["address_country"]],
            region=flat_object[key_map["address_region"]],
            locality=flat_object[key_map["address_locality"]],
            postal_code=flat_object[key_map["address_postalcode"]],
            locality_longitude=flat_object[key_map["address_longitude"]],
            locality_latitude=flat_object[key_map["address_latitude"]]
        )
        incoming_identifier = Identifier(
            scheme=flat_object[key_map["id_schema"]],
            legal_name=flat_object[key_map["id_legal_name"]],
            uri=flat_object[key_map["id_uri"]]
        )
        incoming_contact = ContactPoint(
            name=flat_object[key_map["contact_name"]],
            email=flat_object[key_map["contact_mail"]],
            telephone=flat_object[key_map["contact_phone"]],
            fax_number=flat_object[key_map["contact_fax"]],
            url=flat_object[key_map["contact_url"]],
        )
        try:
            entity = Entity.objects.select_related('identifier', 'address', 'contact_point').get(pk=flat_object[key_map["party_id"]])
            related_fields = [
                ('address', incoming_address),
                ('identifier', incoming_identifier),
                ('contact_point', incoming_contact),
            ]
            set_related_fields(entity, related_fields)
        except Entity.DoesNotExist:
            incoming_address.save()
            incoming_identifier.save()
            incoming_contact.save()
            entity = Entity(
                pk=flat_object[key_map["party_id"]],
                address=incoming_address,
                identifier=incoming_identifier,
                contact_point=incoming_contact
            )
        entity.name = flat_object[key_map["name"]]
        entity.save()
        release.parties.add(entity)

def create_plannings(ws : worksheet):
    key_map = {
        "record_ocid": 0,
        "planning_id": 1,
        "title": 2,
        "description": 3,
        "rationale": 4,
        "budget_source": 5,
        "budget_amount": 6,
        "budget_currency": 7,
        "budget_uri": 8
    }
    releases = get_engaged_objects(Release, ws, 1, 'ref_record__ocid')
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            release = releases[flat_object[key_map["record_ocid"]]]
        except:
            print("Release reference not found")
            continue
        incoming_amount = Value.objects.create(
            amount=flat_object[key_map["budget_amount"]],
            currency=flat_object[key_map["budget_currency"]]
        )
        incoming_project = Projet.objects.create(
            titre_projet=flat_object[key_map["title"]],
            description=flat_object[key_map["description"]]
        )
        incoming_budget = Budget(
            source=flat_object[key_map["budget_source"]],
            uri=flat_object[key_map["budget_uri"]],
            projet=incoming_project,
            amount=incoming_amount,
            description=flat_object[key_map["rationale"]]
        )
        try:
            planning = Planning.objects.select_related('budget').get(pk=flat_object[key_map["planning_id"]])
            related_fields = [
                ('budget', incoming_budget)
            ]
            set_related_fields(planning, related_fields)
        except Planning.DoesNotExist:
            incoming_budget.save()
            planning = Planning(
                pk=flat_object[key_map["planning_id"]],
                budget=incoming_budget
            )
        planning.raison = flat_object[key_map["rationale"]]
        planning.save()
        release.planning = planning
        release.save()

def create_tenders(ws: worksheet):
    key_map = {
        "record_ocid": 0,
        "tender_id": 1,
        "buyer": 2,
        "procuring_entity": 3,
        "title": 4,
        "description": 5,
        "status": 6,
        "procurement_method": 7,
        "procurement_method_rationale": 8,
        "award_criteria": 9,
        "award_criteria_details": 10,
        "submission_method": 11,
        "submission_method_details": 12,
        "min_value_amount": 13,
        "min_value_currency": 14,
        "value_amount": 15,
        "value_currency": 16,
        "tender_period_start": 17,
        "tender_period_end": 18,
        "enquiry_period_start": 19,
        "enquiry_period_end": 20,
        "award_period_start": 21,
        "award_period_end": 22,
        "eligibility_criteria": 23,
    }
    releases = get_engaged_objects(Release, ws, 1, 'ref_record__ocid')
    buyers = get_engaged_objects(Entity, ws, 3, 'pk')
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            release = releases[flat_object[key_map["record_ocid"]]]
            buyer_entity = buyers[flat_object[key_map["buyer"]]]
        except:
            print("Reference not found")
            continue
        try:
            procuring_entity = Entity.objects.get(pk=flat_object[key_map["procuring_entity"]])
        except Entity.DoesNotExist:
            print("Entity with id : %s does not exist" % flat_object[key_map["procuring_entity"]])
            continue

        incoming_min_value = Value(
            amount=flat_object[key_map["min_value_amount"]],
            currency=flat_object[key_map["min_value_currency"]]
        )
        incoming_value = Value(
            amount=flat_object[key_map["value_amount"]],
            currency=flat_object[key_map["value_currency"]]
        )
        incoming_tender_period = Period(
            start_date=flat_object[key_map["tender_period_start"]],
            end_date=flat_object[key_map["tender_period_end"]]
        )
        incoming_enquiry_period = Period(
            start_date=flat_object[key_map["enquiry_period_start"]],
            end_date=flat_object[key_map["enquiry_period_end"]]
        )
        incoming_award_period = Period(
            start_date=flat_object[key_map["award_period_start"]],
            end_date=flat_object[key_map["award_period_end"]]
        )

        try:
            tender = Tender.objects.select_related('min_value', 'value', 'tender_period', 'enquiry_period', 'award_period').get(pk=flat_object[key_map["tender_id"]])
            related_fields = [
                ('min_value', incoming_min_value),
                ('value', incoming_value),
                ('tender_period', incoming_tender_period),
                ('enquiry_period', incoming_enquiry_period),
                ('award_period', incoming_award_period),
            ]
            set_related_fields(tender, related_fields)
        except Tender.DoesNotExist:
            incoming_min_value.save()
            incoming_value.save()
            incoming_tender_period.save()
            incoming_enquiry_period.save()
            incoming_award_period.save()
            tender = Tender(
                pk=flat_object[key_map["tender_id"]],
                min_value=incoming_min_value,
                value=incoming_value,
                tender_period=incoming_tender_period,
                enquiry_period=incoming_enquiry_period,
                award_period=incoming_award_period
            )
        tender.buyer=buyer_entity
        tender.title=flat_object[key_map["title"]]
        tender.description=flat_object[key_map["description"]]
        tender.status=flat_object[key_map["status"]]
        tender.procurement_method=flat_object[key_map["procurement_method"]]
        tender.procurement_method_rationale=flat_object[key_map["procurement_method_rationale"]]
        tender.award_criteria=flat_object[key_map["award_criteria"]]
        tender.award_criteria_details=flat_object[key_map["award_criteria_details"]]
        tender.submission_method=flat_object[key_map["submission_method"]]
        tender.submission_method_details=flat_object[key_map["submission_method_details"]]
        tender.eligibility_criteria=flat_object[key_map["eligibility_criteria"]]
        tender.procuring_entity=procuring_entity
        tender.save()

        release.tender = tender
        release.buyer = buyer_entity
        release.update_buyer_role()
        release.save()

def create_tenderers(ws: worksheet):
    key_map = {
        "tender_id": 0,
        "tenderer_id": 1
    }
    tenders = get_engaged_objects(Tender, ws, 1, 'pk', 'release')
    tenderers = get_engaged_objects(Entity, ws, 2, 'pk')
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            tender = tenders[flat_object[key_map["tender_id"]]]
            tender_release = tender.release
            incoming_tenderer = tenderers[flat_object[key_map["tenderer_id"]]]
        except Release.DoesNotExist:
            print("Release object does not refer to a release")
            continue
        except:
            print("Reference not found")
            continue
        tender.tenderers.add(incoming_tenderer)
        tender_release.parties.add(incoming_tenderer)
        tender_release.add_role(incoming_tenderer, 'tenderer')

def create_awards(ws: worksheet):
    key_map = {
        "record_ocid": 0,
        "id": 1,
        "title": 2,
        "description": 3,
        "status": 4,
        "date": 5,
        "value_amount": 6,
        "value_currency": 7,
        "contract_title": 8,
        "contract_description": 9,
        "contract_status": 10,
        "contract_date": 11,
        "contract_period_start": 12,
        "contract_period_end": 13
    }
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        releases = get_engaged_objects(Release, ws, 1, 'ref_record__ocid')
        try:
            release = releases[flat_object[key_map["record_ocid"]]]
        except:
            print("Release reference not found")
            continue
        incoming_value = Value(
            amount=flat_object[key_map["value_amount"]],
            currency=flat_object[key_map["value_currency"]]
        )
        incoming_contract_period = Period(
            start_date=flat_object[key_map["contract_period_start"]],
            end_date=flat_object[key_map["contract_period_end"]]
        )
        try:
            award = Award.objects.select_related('value', 'contract_period', 'contract__period', 'contract__value', 'release').get(pk=flat_object[key_map["id"]])
            related_fields = [
                ('value', incoming_value),
                ('contract_period', incoming_contract_period)
            ]
            set_related_fields(award, related_fields)
        except Award.DoesNotExist:
            incoming_value.save()
            incoming_contract_period.save()
            award = Award.objects.create(
                pk=flat_object[key_map["id"]],
                value=incoming_value,
                contract_period=incoming_contract_period
            )
        try:
            contract = award.contract
            contract.value = award.value
            contract.period = award.contract_period
        except Contract.DoesNotExist:
            incoming_value.save()
            incoming_contract_period.save()
            contract = Contract.objects.create(
                ref_award=award,
                period=incoming_contract_period,
                value=incoming_value
            )
        try:
            implementation = contract.implementation
        except Implementation.DoesNotExist:
            implementation = Implementation(
                contract = contract
            )
        award.title = flat_object[key_map["title"]]
        award.description = flat_object[key_map["description"]]
        award.status = flat_object[key_map["status"]]
        award.date = flat_object[key_map["date"]]
        award.release = release
        award.save()
        contract.title = flat_object[key_map["contract_title"]]
        contract.description = flat_object[key_map["contract_description"]]
        contract.status = flat_object[key_map["contract_status"]]
        contract.date_signed = flat_object[key_map["contract_date"]]
        contract.save()
        implementation.save()

def create_suppliers(ws: worksheet):
    key_map = {
        "award_id": 0,
        "supplier_id": 1
    }
    awards = get_engaged_objects(Award, ws, 1, 'pk', 'release')
    suppliers = get_engaged_objects(Entity, ws, 2, 'pk')
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            award = awards[flat_object[key_map["award_id"]]]
            award_release = award.release
            incoming_supplier = suppliers[flat_object[key_map["supplier_id"]]]
        except Release.DoesNotExist:
            print("Award object with ID : %s does not refer to a release" % flat_object[key_map["award_id"]])
            continue
        except:
            print("Reference not found")
            continue
        award.suppliers.add(incoming_supplier)
        award_release.parties.add(incoming_supplier)
        award_release.add_role(incoming_supplier, 'supplier')

def create_transactions(ws : worksheet):
    key_map = {
        "award_id": 0,
        "transaction_id": 1,
        "source": 2,
        "uri": 3,
        "date": 4,
        "value_amount": 5,
        "value_currency": 6,
        "payer_id": 7,
        "payee_id": 8
    }
    implementations = get_engaged_objects(Implementation, ws, 1, 'contract__ref_award__pk')
    payers = get_engaged_objects(Entity, ws, 8, 'pk')
    payees = get_engaged_objects(Entity, ws, 9, 'pk')
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            implementation = implementations[flat_object[key_map["award_id"]]]
            payer = payers[flat_object[key_map["payer_id"]]]
            payee = payees[flat_object[key_map["payee_id"]]]
        except:
            print("Reference not found")
            continue
        incoming_value = Value(
            amount=flat_object[key_map["value_amount"]],
            currency=flat_object[key_map["value_currency"]]
        )
        try:
            transaction = Transaction.objects.select_related('value', 'implementation', 'payer', 'payee').get(pk=flat_object[key_map["transaction_id"]])
            related_fields = [
                ('value', incoming_value)
            ]
            set_related_fields(transaction, related_fields)
        except Transaction.DoesNotExist:
            incoming_value.save()
            transaction = Transaction(
                pk=flat_object[key_map["transaction_id"]],
                value=incoming_value
            )
        transaction.implementation = implementation
        transaction.source = flat_object[key_map["source"]]
        transaction.uri = flat_object[key_map["uri"]]
        transaction.date = flat_object[key_map["date"]]
        transaction.payer = payer
        transaction.payee = payee
        transaction.save()

def create_items(ws : worksheet):
    key_map = {
        "item_id": 0,
        "step_name": 1,
        "ref_step": 2,
        "description": 3,
        "quantity": 4,
        "unit_name": 5,
        "unit_value_amount": 6,
        "unit_value_currency": 7,
        "classification_scheme": 8,
        "classification_description": 9,
        "classification_uri": 10
    }
    modifiers = {
        'tender': Tender,
        'award': Award,
        'contract': Contract
    }
    ref_objects = get_multiple_engaged_objects(ws, 3, 'pk', modifiers, 2, prefetch_related_fields=('items'))
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            step_name = flat_object[key_map["step_name"]]
            ref_object = ref_objects[step_name][flat_object[key_map["ref_step"]]]
        except:
            print("Reference not found")
        items = ref_object.items
        incoming_value = Value(
            amount=flat_object[key_map["unit_value_amount"]],
            currency=flat_object[key_map["unit_value_currency"]]
        )
        incoming_unit = Unit(
            name=flat_object[key_map["unit_name"]],
            value=incoming_value
        )
        incoming_classification = Classification(
            scheme=flat_object[key_map["classification_scheme"]],
            description=flat_object[key_map["classification_description"]],
            uri=flat_object[key_map["classification_uri"]]
        )
        try:
            item = items.select_related('classification', 'unit').get(pk=flat_object[key_map["item_id"]])
            related_fields = [
                ('classification', incoming_classification),
                ('unit', incoming_unit),
            ]
            set_related_fields(item, related_fields)
            item.quantity = flat_object[key_map["quantity"]]
            item.description = flat_object[key_map["description"]]
            item.save()
        except:
            incoming_value.save()
            incoming_unit.save()
            incoming_classification.save()
            item = items.create(
                pk=flat_object[key_map["item_id"]],
                quantity = flat_object[key_map["quantity"]],
                description = flat_object[key_map["description"]],
                classification=incoming_classification,
                unit=incoming_unit
            )

def create_milestones(ws : worksheet):
    key_map = {
        "milestone_id": 0,
        "step_name": 1,
        "ref_step": 2,
        "title": 3,
        "description": 4,
        "due_date": 5,
        "date_modified": 6,
        "status": 7
    }
    modifiers = {
        'planning': Planning,
        'tender': Tender,
        'contract': Contract,
        'implementation': Implementation,
    }
    ref_objects = get_multiple_engaged_objects(ws, 3, 'pk', modifiers, 2, prefetch_related_fields=('milestones'))
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            step_name = flat_object[key_map["step_name"]]
            ref_object = ref_objects[step_name][flat_object[key_map["ref_step"]]]
        except:
            print("Reference not found")
        milestones = ref_object.milestones
        try:
            milestone = milestones.get(pk=flat_object[key_map["milestone_id"]])
            milestone.title = flat_object[key_map["title"]]
            milestone.description = flat_object[key_map["description"]]
            milestone.due_date = flat_object[key_map["due_date"]]
            milestone.date_modified = flat_object[key_map["date_modified"]]
            milestone.status = flat_object[key_map["status"]]
            milestone.save()
        except:
            milestone = milestones.create(
                pk=flat_object[key_map["milestone_id"]],
                title = flat_object[key_map["title"]],
                description = flat_object[key_map["description"]],
                due_date = flat_object[key_map["due_date"]],
                date_modified = flat_object[key_map["date_modified"]],
                status = flat_object[key_map["status"]]
            )

def create_documents(ws: worksheet):
    key_map = {
        "document_id": 0,
        "step_name": 1,
        "ref_step": 2,
        "type": 3,
        "title": 4,
        "description": 5,
        "url": 6,
        "date_published": 7,
        "date_modified": 8,
        "format": 9,
        "language": 10
    }
    modifiers = {
        'planning': Planning,
        'tender': Tender,
        'award': Award,
        'contract': Contract,
        'implementation': Implementation,
        'milestone': Milestone
    }
    ref_objects = get_multiple_engaged_objects(ws, 3, 'pk', modifiers, 2, prefetch_related_fields=('documents'))
    for flat_object in ws.iter_rows(min_row=4, values_only=True):
        if not flat_object[0]:
            continue
        try:
            step_name = flat_object[key_map["step_name"]]
            ref_object = ref_objects[step_name][flat_object[key_map["ref_step"]]]
        except:
            print("%s object with ID : %s does not exist" % (flat_object[key_map["step_name"]], flat_object[key_map["ref_step"]]))
            continue
        documents = ref_object.documents
        try:
            document = documents.get(pk=flat_object[key_map["document_id"]])
            document.document_type = flat_object[key_map["type"]]
            document.title = flat_object[key_map["title"]]
            document.description = flat_object[key_map["description"]]
            document.url = flat_object[key_map["url"]]
            document.date_published = flat_object[key_map["date_published"]]
            document.date_modified = flat_object[key_map["date_modified"]]
            document.document_format = flat_object[key_map["format"]]
            document.language = flat_object[key_map["language"]]
            document.save()
        except:
            document = documents.create(
                pk = flat_object[key_map["document_id"]],
                document_type = flat_object[key_map["type"]],
                title = flat_object[key_map["title"]],
                description = flat_object[key_map["description"]],
                url = flat_object[key_map["url"]],
                date_published = flat_object[key_map["date_published"]],
                date_modified = flat_object[key_map["date_modified"]],
                document_format = flat_object[key_map["format"]],
                language = flat_object[key_map["language"]]
            )
